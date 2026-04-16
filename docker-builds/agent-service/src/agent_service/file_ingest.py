from __future__ import annotations

import base64
import binascii
import csv
import io
import json
from typing import Iterable

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

from agent_service.models import ChatAttachment

MAX_ATTACHMENTS = 8
MAX_BYTES_PER_FILE = 5 * 1024 * 1024
MAX_TOTAL_CHARS = 40000
MAX_CHARS_PER_FILE = 10000


def _truncate(text: str, limit: int) -> str:
    if len(text) <= limit:
        return text
    return text[:limit] + "\n...[truncated]"


def _decode_text(data: bytes) -> str:
    try:
        return data.decode("utf-8")
    except UnicodeDecodeError:
        return data.decode("latin-1", errors="replace")


def _extract_csv_text(data: bytes) -> str:
    text = _decode_text(data)
    lines = text.splitlines()
    reader = csv.reader(lines)
    rows: list[str] = []
    for idx, row in enumerate(reader):
        if idx >= 200:
            rows.append("...[truncated rows]")
            break
        rows.append(" | ".join(cell.strip() for cell in row))
    return "\n".join(rows)


def _extract_xlsx_text(data: bytes) -> str:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sections: list[str] = []
    for sheet_idx, sheet in enumerate(workbook.worksheets):
        if sheet_idx >= 5:
            sections.append("...[truncated sheets]")
            break
        sections.append(f"# Sheet: {sheet.title}")
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx >= 200:
                sections.append("...[truncated rows]")
                break
            values = ["" if value is None else str(value) for value in row]
            sections.append(" | ".join(values).strip())
    return "\n".join(line for line in sections if line)


def _extract_pdf_text(data: bytes) -> str:
    reader = PdfReader(io.BytesIO(data))
    parts: list[str] = []
    for page_idx, page in enumerate(reader.pages):
        if page_idx >= 40:
            parts.append("...[truncated pages]")
            break
        page_text = page.extract_text() or ""
        if page_text.strip():
            parts.append(page_text)
    return "\n\n".join(parts)


def _extract_docx_text(data: bytes) -> str:
    doc = Document(io.BytesIO(data))
    lines = [paragraph.text for paragraph in doc.paragraphs if paragraph.text.strip()]
    return "\n".join(lines)


def _extension(filename: str) -> str:
    if "." not in filename:
        return ""
    return filename.rsplit(".", 1)[-1].lower()


def _parse_attachment_text(attachment: ChatAttachment, raw: bytes) -> str:
    ext = _extension(attachment.filename)

    if ext in {"txt", "md", "log", "yaml", "yml", "ini", "cfg", "py", "js", "ts", "json"}:
        text = _decode_text(raw)
        if ext == "json":
            try:
                parsed = json.loads(text)
                text = json.dumps(parsed, indent=2, ensure_ascii=True)
            except json.JSONDecodeError:
                pass
        return text

    if ext == "csv":
        return _extract_csv_text(raw)

    if ext == "pdf":
        return _extract_pdf_text(raw)

    if ext == "docx":
        return _extract_docx_text(raw)

    if ext == "xlsx":
        return _extract_xlsx_text(raw)

    return _decode_text(raw)


def _iter_attachment_sections(attachments: Iterable[ChatAttachment]) -> Iterable[str]:
    for attachment in attachments:
        try:
            raw = base64.b64decode(attachment.content_base64, validate=True)
        except (binascii.Error, ValueError):
            yield f"## {attachment.filename}\n[Could not decode attachment: invalid base64]"
            continue

        if len(raw) > MAX_BYTES_PER_FILE:
            yield f"## {attachment.filename}\n[Skipped: file exceeds {MAX_BYTES_PER_FILE} bytes limit]"
            continue

        try:
            text = _parse_attachment_text(attachment, raw)
        except Exception as exc:  # noqa: BLE001
            yield f"## {attachment.filename}\n[Could not parse attachment: {exc}]"
            continue

        cleaned = _truncate(text.strip(), MAX_CHARS_PER_FILE)
        if not cleaned:
            cleaned = "[No extractable text found]"

        yield f"## {attachment.filename}\n{cleaned}"


def build_attachments_context(attachments: list[ChatAttachment]) -> str:
    if not attachments:
        return ""

    selected = attachments[:MAX_ATTACHMENTS]
    sections = list(_iter_attachment_sections(selected))
    if not sections:
        return ""

    content = "\n\n".join(sections)
    return _truncate(content, MAX_TOTAL_CHARS)
